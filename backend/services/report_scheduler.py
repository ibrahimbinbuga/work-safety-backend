"""Scheduled automatic report emailer.

Runs daily at 08:00 UTC and checks each company's notification settings.
Sends a report in every requested format (PDF / Excel / CSV) when the
company's chosen period (daily / weekly / monthly) matches today's date.
"""
import csv
import io
import os
import smtplib
import ssl
from datetime import datetime, date, timedelta, time
from email.message import EmailMessage

from sqlalchemy import select, and_
from sqlalchemy.ext.asyncio import AsyncSession

import models
from database import AsyncSessionLocal


# ---------------------------------------------------------------------------
# Helpers – report generation
# ---------------------------------------------------------------------------

def _period_matches(period: str) -> bool:
    today = datetime.utcnow()
    if period == "daily":
        return True
    if period == "weekly":
        return today.weekday() == 0          # Monday
    if period == "monthly":
        return today.day == 1                # 1st of month
    return False


def _date_range_for_period(period: str):
    today = date.today()
    if period == "daily":
        return today - timedelta(days=1), today
    if period == "weekly":
        return today - timedelta(days=7), today
    return (today.replace(day=1) - timedelta(days=1)).replace(day=1), today


async def _fetch_violations(db: AsyncSession, company_id: int, from_date: date, to_date: date):
    result = await db.execute(
        select(models.Violations)
        .where(
            and_(
                models.Violations.company_id == company_id,
                models.Violations.tarih_saat >= datetime.combine(from_date, time.min),
                models.Violations.tarih_saat <= datetime.combine(to_date, time.max),
            )
        )
        .order_by(models.Violations.tarih_saat.desc())
    )
    return result.scalars().all()


VIOLATION_LABELS = {"head": "No Helmet", "vest": "No Vest", "fallen": "Fall Detected"}
HEADERS = ["ID", "Type", "Camera/Zone", "Worker ID", "Timestamp", "Status"]


def _rows(violations):
    return [
        [
            v.id,
            VIOLATION_LABELS.get(v.ihlal_cesidi, v.ihlal_cesidi),
            v.ihlal_yapilan_bolge or "-",
            v.worker_id or "-",
            v.tarih_saat.strftime("%Y-%m-%d %H:%M:%S") if v.tarih_saat else "-",
            v.review_status or "pending",
        ]
        for v in violations
    ]


def _generate_csv(violations) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(HEADERS)
    writer.writerows(_rows(violations))
    return buf.getvalue().encode("utf-8-sig")


def _generate_excel(violations) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Violations"
    header_fill = PatternFill("solid", fgColor="2563EB")
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row_data in _rows(violations):
        ws.append(row_data)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generate_pdf(violations, company_code: str, from_date: date, to_date: date) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=15*mm, rightMargin=15*mm,
                             topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f"Safety Violations Report – {company_code}", styles["Title"]),
        Paragraph(f"Period: {from_date} → {to_date}", styles["Normal"]),
        Spacer(1, 8*mm),
    ]
    data = [HEADERS] + _rows(violations)
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Email sender
# ---------------------------------------------------------------------------

def _send_email(to_addresses: list[str], subject: str, body: str, attachments: list[tuple]):
    smtp_host = os.getenv("SMTP_HOST", "")
    smtp_port = int(os.getenv("SMTP_PORT", 587))
    smtp_user = os.getenv("SMTP_USER", "")
    smtp_password = os.getenv("SMTP_PASSWORD", "")
    smtp_from = os.getenv("SMTP_FROM", smtp_user)

    if not smtp_host or not smtp_user:
        print("[Scheduler] SMTP not configured – skipping email")
        return

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = smtp_from
    msg["To"] = ", ".join(to_addresses)
    msg.set_content(body)

    for filename, data, mime_main, mime_sub in attachments:
        msg.add_attachment(data, maintype=mime_main, subtype=mime_sub, filename=filename)

    with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as server:
        server.starttls(context=ssl.create_default_context())
        server.login(smtp_user, smtp_password)
        server.send_message(msg)


# ---------------------------------------------------------------------------
# Main scheduled job
# ---------------------------------------------------------------------------

async def send_scheduled_reports():
    print(f"[Scheduler] Running report job at {datetime.utcnow().isoformat()}")
    async with AsyncSessionLocal() as db:
        # Load all notification settings with email enabled
        result = await db.execute(
            select(models.CompanyNotificationSettings)
            .where(models.CompanyNotificationSettings.email_enabled == True)
        )
        all_settings = result.scalars().all()

        for ns in all_settings:
            if not _period_matches(ns.report_period):
                continue

            # Get company
            company = await db.get(models.Company, ns.company_id)
            if not company:
                continue

            from_date, to_date = _date_range_for_period(ns.report_period)

            # Get recipient users (all active users of this company)
            users_result = await db.execute(
                select(models.User).where(
                    models.User.company_id == ns.company_id,
                    models.User.is_active == True,
                )
            )
            recipients = [u.email for u in users_result.scalars().all()]
            if not recipients:
                print(f"[Scheduler] No active users for {company.code}, skipping")
                continue

            violations = await _fetch_violations(db, ns.company_id, from_date, to_date)
            formats = ns.report_formats or ["pdf"]
            attachments = []

            if "csv" in formats:
                attachments.append((
                    f"violations_{company.code}_{to_date}.csv",
                    _generate_csv(violations),
                    "text", "csv",
                ))
            if "excel" in formats:
                attachments.append((
                    f"violations_{company.code}_{to_date}.xlsx",
                    _generate_excel(violations),
                    "application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ))
            if "pdf" in formats:
                attachments.append((
                    f"violations_{company.code}_{to_date}.pdf",
                    _generate_pdf(violations, company.code, from_date, to_date),
                    "application", "pdf",
                ))

            period_label = {"daily": "Daily", "weekly": "Weekly", "monthly": "Monthly"}[ns.report_period]
            subject = f"SafetyWatch {period_label} Report – {company.code}"
            body = (
                f"Hello,\n\n"
                f"Please find attached the {period_label.lower()} safety violations report "
                f"for {company.name} ({company.code}).\n"
                f"Period: {from_date} → {to_date}\n"
                f"Total violations: {len(violations)}\n\n"
                f"Best regards,\nSafetyWatch"
            )

            try:
                _send_email(recipients, subject, body, attachments)
                print(f"[Scheduler] Sent {period_label} report for {company.code} to {recipients}")
            except Exception as exc:
                print(f"[Scheduler] Failed to send report for {company.code}: {exc}")
