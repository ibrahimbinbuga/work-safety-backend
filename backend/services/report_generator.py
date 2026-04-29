"""Shared report generation utilities.

Used by both routes/reports.py (on-demand export) and
services/report_scheduler.py (scheduled email attachments).
"""
import csv
import io
from collections import defaultdict
from datetime import date, datetime

VIOLATION_TYPE_LABELS = {
    "head": "No Helmet",
    "vest": "No Vest",
    "fallen": "Fall Detected",
}

VIOLATION_TYPE_MODELS = {
    "head": "PPE Model",
    "vest": "PPE Model",
    "fallen": "Fall Detection Model",
}


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

def generate_csv(violations, company_code: str, from_date: date, to_date: date) -> bytes:
    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=["ID", "Type", "Model", "Camera", "Date & Time", "Status", "Worker ID"],
    )
    writer.writeheader()
    for v in violations:
        writer.writerow({
            "ID": v.id,
            "Type": VIOLATION_TYPE_LABELS.get(v.ihlal_cesidi, v.ihlal_cesidi),
            "Model": VIOLATION_TYPE_MODELS.get(v.ihlal_cesidi, "Unknown"),
            "Camera": f"Camera {v.ihlal_yapilan_bolge}" if v.ihlal_yapilan_bolge else "Unknown",
            "Date & Time": v.tarih_saat.strftime("%Y-%m-%d %H:%M:%S") if v.tarih_saat else "",
            "Status": (v.review_status or "pending").capitalize(),
            "Worker ID": v.violation_id,
        })
    return output.getvalue().encode("utf-8-sig")


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def generate_excel(
    violations,
    company_code: str,
    company_name: str,
    from_date: date,
    to_date: date,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Violations Report"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="1E3A5F")
    meta_font = Font(size=10, color="6B7280")
    border_side = Side(style="thin", color="E5E7EB")
    cell_border = Border(
        left=border_side, right=border_side, top=border_side, bottom=border_side
    )
    center = Alignment(horizontal="center", vertical="center")

    TYPE_COLORS = {"head": "FEE2E2", "vest": "FFEDD5", "fallen": "F3E8FF"}
    STATUS_COLORS = {"pending": "FEF9C3", "reviewed": "DBEAFE", "resolved": "DCFCE7"}

    # Title block
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Safety Violations Report — {company_name or company_code.upper()}"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Period: {from_date}  →  {to_date}     Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = meta_font
    ws["A2"].alignment = center

    ws.merge_cells("A3:G3")
    ws["A3"] = f"Total records: {len(violations)}"
    ws["A3"].font = meta_font
    ws["A3"].alignment = center

    # Header row
    headers = ["ID", "Type", "Model", "Camera", "Date & Time", "Status", "Worker ID"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = cell_border

    # Data rows
    for row_idx, v in enumerate(violations, start=6):
        row_data = [
            v.id,
            VIOLATION_TYPE_LABELS.get(v.ihlal_cesidi, v.ihlal_cesidi),
            VIOLATION_TYPE_MODELS.get(v.ihlal_cesidi, "Unknown"),
            f"Camera {v.ihlal_yapilan_bolge}" if v.ihlal_yapilan_bolge else "Unknown",
            v.tarih_saat.strftime("%Y-%m-%d %H:%M:%S") if v.tarih_saat else "",
            (v.review_status or "pending").capitalize(),
            v.violation_id,
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center")
            if col_idx == 2:
                cell.fill = PatternFill("solid", fgColor=TYPE_COLORS.get(v.ihlal_cesidi, "FFFFFF"))
            elif col_idx == 6:
                cell.fill = PatternFill("solid", fgColor=STATUS_COLORS.get(v.review_status or "pending", "FFFFFF"))
            elif row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F9FAFB")

    # Column widths
    for i, width in enumerate([8, 18, 22, 18, 22, 14, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[5].height = 22

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Violation Type"
    ws2["B1"] = "Count"
    ws2["A1"].font = header_font
    ws2["B1"].font = header_font
    ws2["A1"].fill = header_fill
    ws2["B1"].fill = header_fill

    by_type: dict[str, int] = defaultdict(int)
    by_status: dict[str, int] = defaultdict(int)
    for v in violations:
        by_type[v.ihlal_cesidi] += 1
        by_status[v.review_status or "pending"] += 1

    for row_idx, (t, cnt) in enumerate(sorted(by_type.items()), start=2):
        ws2.cell(row=row_idx, column=1, value=VIOLATION_TYPE_LABELS.get(t, t))
        ws2.cell(row=row_idx, column=2, value=cnt)

    offset = len(by_type) + 4
    ws2.cell(row=offset, column=1, value="Review Status").font = header_font
    ws2.cell(row=offset, column=2, value="Count").font = header_font
    ws2.cell(row=offset, column=1).fill = header_fill
    ws2.cell(row=offset, column=2).fill = header_fill
    for i, (s, cnt) in enumerate(sorted(by_status.items()), start=offset + 1):
        ws2.cell(row=i, column=1, value=s.capitalize())
        ws2.cell(row=i, column=2, value=cnt)

    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 12

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def generate_pdf(
    violations,
    company_code: str,
    company_name: str,
    from_date: date,
    to_date: date,
) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title", parent=styles["Title"],
        fontSize=16, textColor=colors.HexColor("#1E3A5F"),
        alignment=TA_CENTER,
    )
    meta_style = ParagraphStyle(
        "meta", parent=styles["Normal"],
        fontSize=9, textColor=colors.HexColor("#6B7280"),
        alignment=TA_CENTER,
    )

    label = company_name or company_code.upper()
    elements = [
        Paragraph(f"Safety Violations Report — {label}", title_style),
        Paragraph(
            f"Period: {from_date} → {to_date} &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"Total: {len(violations)}",
            meta_style,
        ),
        Spacer(1, 8 * mm),
    ]

    headers = ["ID", "Type", "Model", "Camera", "Date & Time", "Status", "Worker ID"]
    rows = [headers]
    for v in violations:
        rows.append([
            str(v.id),
            VIOLATION_TYPE_LABELS.get(v.ihlal_cesidi, v.ihlal_cesidi),
            VIOLATION_TYPE_MODELS.get(v.ihlal_cesidi, "Unknown"),
            f"Camera {v.ihlal_yapilan_bolge}" if v.ihlal_yapilan_bolge else "Unknown",
            v.tarih_saat.strftime("%Y-%m-%d %H:%M:%S") if v.tarih_saat else "-",
            (v.review_status or "pending").capitalize(),
            str(v.violation_id) if v.violation_id else "-",
        ])

    col_widths = [18*mm, 28*mm, 40*mm, 35*mm, 45*mm, 25*mm, 22*mm]
    table = Table(rows, colWidths=col_widths, repeatRows=1)

    TYPE_PDF_COLORS = {
        "head": colors.HexColor("#FEE2E2"),
        "vest": colors.HexColor("#FFEDD5"),
        "fallen": colors.HexColor("#F3E8FF"),
    }
    STATUS_PDF_COLORS = {
        "pending": colors.HexColor("#FEF9C3"),
        "reviewed": colors.HexColor("#DBEAFE"),
        "resolved": colors.HexColor("#DCFCE7"),
    }

    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E7EB")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
    ]
    for row_idx, v in enumerate(violations, start=1):
        type_color = TYPE_PDF_COLORS.get(v.ihlal_cesidi)
        if type_color:
            style_cmds.append(("BACKGROUND", (1, row_idx), (1, row_idx), type_color))
        status_color = STATUS_PDF_COLORS.get(v.review_status or "pending")
        if status_color:
            style_cmds.append(("BACKGROUND", (5, row_idx), (5, row_idx), status_color))

    table.setStyle(TableStyle(style_cmds))
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()
