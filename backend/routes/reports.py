# backend/routes/reports.py
"""Reporting endpoints: summary data + CSV/Excel exports."""
import csv
import io
from collections import defaultdict
from datetime import datetime, date, timedelta, time
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select, and_
from sqlalchemy.ext.asyncio import AsyncSession

from auth import TokenData
from database import get_db
from dependencies import get_current_user, verify_company_access
import models

router = APIRouter()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

VIOLATION_TYPE_LABELS = {
    "head": "No Helmet",
    "vest": "No Vest",
    "fallen": "Fall Detected",
}

VIOLATION_TYPE_MODELS = {
    "head": "PPE Model",
    "vest": "PPE Model",
    "fallen": "Fall Detection Model",
}


async def _get_company(db: AsyncSession, company_code: str) -> models.Company:
    result = await db.execute(
        select(models.Company).where(
            func.upper(models.Company.code) == func.upper(company_code)
        )
    )
    company = result.scalar_one_or_none()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    return company


def _build_filters(
    company_id: int,
    from_date: Optional[date],
    to_date: Optional[date],
    violation_types: List[str],
    review_statuses: List[str],
):
    filters = [models.Violations.company_id == company_id]
    if from_date:
        filters.append(
            models.Violations.tarih_saat >= datetime.combine(from_date, time.min)
        )
    if to_date:
        filters.append(
            models.Violations.tarih_saat <= datetime.combine(to_date, time.max)
        )
    if violation_types:
        filters.append(models.Violations.ihlal_cesidi.in_(violation_types))
    if review_statuses:
        filters.append(models.Violations.review_status.in_(review_statuses))
    return filters


# ---------------------------------------------------------------------------
# Active violation types for a company (based on assigned models)
# ---------------------------------------------------------------------------

# Maps violation types to their model group label
_TYPE_TO_MODEL_GROUP = {
    "head":   "PPE Model",
    "vest":   "PPE Model",
    "fallen": "Fall Detection Model",
}


def _model_path_to_types(path: str) -> list[str]:
    """Infer applicable violation types from a model file path.

    Convention used in this project:
      - fall_model/... or any path containing 'fall' → Fall Detection
      - everything else                              → PPE (head + vest)
    """
    if "fall" in path.lower():
        return ["fallen"]
    return ["head", "vest"]


@router.get("/api/company/{company_code}/reports/active-types")
async def get_active_violation_types(
    company_code: str,
    current_user: TokenData = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Return the violation types that are relevant for this company
    based on which models are currently assigned (is_active=True)."""
    await verify_company_access(current_user, company_code)
    company = await _get_company(db, company_code)

    result = await db.execute(
        select(models.ModelMeta)
        .join(models.CompanyModel, models.CompanyModel.model_id == models.ModelMeta.id)
        .where(
            models.CompanyModel.company_id == company.id,
            models.CompanyModel.is_active == True,
        )
    )
    assigned_models = result.scalars().all()

    active_types: list[str] = []
    for m in assigned_models:
        for t in _model_path_to_types(m.path or ""):
            if t not in active_types:
                active_types.append(t)

    # Fallback: if no models assigned yet, return all types so UI isn't empty
    if not active_types:
        active_types = ["head", "vest", "fallen"]

    return {
        "types": active_types,
        "groups": [
            {
                "label": group_label,
                "types": [t for t in active_types if _TYPE_TO_MODEL_GROUP.get(t) == group_label],
            }
            for group_label in dict.fromkeys(
                _TYPE_TO_MODEL_GROUP[t] for t in active_types if t in _TYPE_TO_MODEL_GROUP
            )
        ],
    }


# ---------------------------------------------------------------------------
# Report data endpoint
# ---------------------------------------------------------------------------

@router.get("/api/company/{company_code}/reports/data")
async def get_report_data(
    company_code: str,
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    violation_types: List[str] = Query(default=[]),
    review_statuses: List[str] = Query(default=[]),
    current_user: TokenData = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await verify_company_access(current_user, company_code)
    company = await _get_company(db, company_code)

    filters = _build_filters(company.id, from_date, to_date, violation_types, review_statuses)

    result = await db.execute(
        select(models.Violations)
        .where(and_(*filters))
        .order_by(models.Violations.tarih_saat.asc())
    )
    violations = result.scalars().all()

    # --- aggregations ---
    by_type: dict[str, int] = defaultdict(int)
    by_status: dict[str, int] = defaultdict(int)
    by_camera: dict[str, int] = defaultdict(int)
    by_day: dict[str, int] = defaultdict(int)

    for v in violations:
        by_type[v.ihlal_cesidi] += 1
        by_status[v.review_status or "pending"] += 1
        cam_label = f"Camera {v.ihlal_yapilan_bolge}" if v.ihlal_yapilan_bolge else "Unknown"
        by_camera[cam_label] += 1
        if v.tarih_saat:
            day_key = v.tarih_saat.strftime("%Y-%m-%d")
            by_day[day_key] += 1

    # Top camera
    top_camera = max(by_camera, key=by_camera.get) if by_camera else "—"

    # Fill missing days in range so chart is continuous
    daily_data = []
    if from_date and to_date:
        current = from_date
        while current <= to_date:
            key = current.strftime("%Y-%m-%d")
            daily_data.append({
                "date": current.strftime("%b %d"),
                "full_date": key,
                "violations": by_day.get(key, 0),
            })
            current += timedelta(days=1)
    else:
        for key in sorted(by_day.keys()):
            dt = datetime.strptime(key, "%Y-%m-%d")
            daily_data.append({
                "date": dt.strftime("%b %d"),
                "full_date": key,
                "violations": by_day[key],
            })

    # Violation distribution for pie chart
    violation_distribution = [
        {
            "name": VIOLATION_TYPE_LABELS.get(t, t),
            "value": cnt,
            "type": t,
        }
        for t, cnt in sorted(by_type.items(), key=lambda x: -x[1])
    ]

    # Camera bar chart data
    camera_data = [
        {"camera": cam, "violations": cnt}
        for cam, cnt in sorted(by_camera.items(), key=lambda x: -x[1])[:10]
    ]

    return {
        "summary": {
            "total": len(violations),
            "by_type": dict(by_type),
            "by_status": dict(by_status),
            "top_camera": top_camera,
            "pending": by_status.get("pending", 0),
            "reviewed": by_status.get("reviewed", 0),
            "resolved": by_status.get("resolved", 0),
        },
        "daily_data": daily_data,
        "violation_distribution": violation_distribution,
        "camera_data": camera_data,
        "violations": [
            {
                "id": v.id,
                "type": v.ihlal_cesidi,
                "type_label": VIOLATION_TYPE_LABELS.get(v.ihlal_cesidi, v.ihlal_cesidi),
                "model": VIOLATION_TYPE_MODELS.get(v.ihlal_cesidi, "Unknown"),
                "camera_id": v.ihlal_yapilan_bolge,
                "camera_label": f"Camera {v.ihlal_yapilan_bolge}" if v.ihlal_yapilan_bolge else "Unknown",
                "datetime": v.tarih_saat.isoformat() if v.tarih_saat else None,
                "review_status": v.review_status or "pending",
                "worker_id": v.violation_id,
            }
            for v in reversed(violations)  # newest first
        ],
    }


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------

@router.get("/api/company/{company_code}/reports/export/csv")
async def export_csv(
    company_code: str,
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    violation_types: List[str] = Query(default=[]),
    review_statuses: List[str] = Query(default=[]),
    current_user: TokenData = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await verify_company_access(current_user, company_code)
    company = await _get_company(db, company_code)

    filters = _build_filters(company.id, from_date, to_date, violation_types, review_statuses)
    result = await db.execute(
        select(models.Violations)
        .where(and_(*filters))
        .order_by(models.Violations.tarih_saat.desc())
    )
    violations = result.scalars().all()

    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=["ID", "Type", "Model", "Camera", "Date & Time", "Status", "Worker ID"],
    )
    writer.writeheader()
    for v in violations:
        writer.writerow({
            "ID": v.id,
            "Type": VIOLATION_TYPE_LABELS.get(v.ihlal_cesidi, v.ihlal_cesidi),
            "Model": VIOLATION_TYPE_MODELS.get(v.ihlal_cesidi, "Unknown"),
            "Camera": f"Camera {v.ihlal_yapilan_bolge}" if v.ihlal_yapilan_bolge else "Unknown",
            "Date & Time": v.tarih_saat.strftime("%Y-%m-%d %H:%M:%S") if v.tarih_saat else "",
            "Status": (v.review_status or "pending").capitalize(),
            "Worker ID": v.violation_id,
        })

    filename = f"violations_{company_code}_{date.today().isoformat()}.csv"
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue().encode("utf-8-sig")]),  # utf-8-sig for Excel compatibility
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

@router.get("/api/company/{company_code}/reports/export/excel")
async def export_excel(
    company_code: str,
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    violation_types: List[str] = Query(default=[]),
    review_statuses: List[str] = Query(default=[]),
    current_user: TokenData = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is not installed. Run: pip install openpyxl",
        )

    await verify_company_access(current_user, company_code)
    company = await _get_company(db, company_code)

    filters = _build_filters(company.id, from_date, to_date, violation_types, review_statuses)
    result = await db.execute(
        select(models.Violations)
        .where(and_(*filters))
        .order_by(models.Violations.tarih_saat.desc())
    )
    violations = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Violations Report"

    # ---- Styles ----
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="1E3A5F")
    meta_font = Font(size=10, color="6B7280")
    border_side = Side(style="thin", color="E5E7EB")
    cell_border = Border(
        left=border_side, right=border_side, top=border_side, bottom=border_side
    )
    center = Alignment(horizontal="center", vertical="center")

    TYPE_COLORS = {
        "head": "FEE2E2",
        "vest": "FFEDD5",
        "fallen": "F3E8FF",
    }
    STATUS_COLORS = {
        "pending": "FEF9C3",
        "reviewed": "DBEAFE",
        "resolved": "DCFCE7",
    }

    # ---- Title block ----
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Safety Violations Report — {company.name if hasattr(company, 'name') else company_code.upper()}"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws.merge_cells("A2:G2")
    period_str = ""
    if from_date and to_date:
        period_str = f"{from_date}  →  {to_date}"
    ws["A2"] = f"Period: {period_str}     Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = meta_font
    ws["A2"].alignment = center

    ws.merge_cells("A3:G3")
    ws["A3"] = f"Total records: {len(violations)}"
    ws["A3"].font = meta_font
    ws["A3"].alignment = center

    # ---- Header row ----
    headers = ["ID", "Type", "Model", "Camera", "Date & Time", "Status", "Worker ID"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = cell_border

    # ---- Data rows ----
    for row_idx, v in enumerate(violations, start=6):
        row_data = [
            v.id,
            VIOLATION_TYPE_LABELS.get(v.ihlal_cesidi, v.ihlal_cesidi),
            VIOLATION_TYPE_MODELS.get(v.ihlal_cesidi, "Unknown"),
            f"Camera {v.ihlal_yapilan_bolge}" if v.ihlal_yapilan_bolge else "Unknown",
            v.tarih_saat.strftime("%Y-%m-%d %H:%M:%S") if v.tarih_saat else "",
            (v.review_status or "pending").capitalize(),
            v.violation_id,
        ]
        type_color = TYPE_COLORS.get(v.ihlal_cesidi, "FFFFFF")
        status_color = STATUS_COLORS.get(v.review_status or "pending", "FFFFFF")

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center")
            if col_idx == 2:  # Type column
                cell.fill = PatternFill("solid", fgColor=type_color)
            elif col_idx == 6:  # Status column
                cell.fill = PatternFill("solid", fgColor=status_color)
            if row_idx % 2 == 0 and col_idx not in (2, 6):
                cell.fill = PatternFill("solid", fgColor="F9FAFB")

    # ---- Column widths ----
    col_widths = [8, 18, 22, 18, 22, 14, 12]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[5].height = 22

    # ---- Summary sheet ----
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Violation Type"
    ws2["B1"] = "Count"
    ws2["A1"].font = header_font
    ws2["B1"].font = header_font
    ws2["A1"].fill = header_fill
    ws2["B1"].fill = header_fill

    by_type: dict[str, int] = defaultdict(int)
    by_status: dict[str, int] = defaultdict(int)
    for v in violations:
        by_type[v.ihlal_cesidi] += 1
        by_status[v.review_status or "pending"] += 1

    for row_idx, (t, cnt) in enumerate(sorted(by_type.items()), start=2):
        ws2.cell(row=row_idx, column=1, value=VIOLATION_TYPE_LABELS.get(t, t))
        ws2.cell(row=row_idx, column=2, value=cnt)

    ws2["A" + str(len(by_type) + 4)] = "Review Status"
    ws2["B" + str(len(by_type) + 4)] = "Count"
    ws2["A" + str(len(by_type) + 4)].font = header_font
    ws2["B" + str(len(by_type) + 4)].font = header_font
    for i, (s, cnt) in enumerate(sorted(by_status.items()), start=len(by_type) + 5):
        ws2.cell(row=i, column=1, value=s.capitalize())
        ws2.cell(row=i, column=2, value=cnt)

    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 12

    # ---- Stream ----
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"violations_{company_code}_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
